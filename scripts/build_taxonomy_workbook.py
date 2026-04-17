#!/usr/bin/env python3
"""Build Course_Taxonomy_Master.xlsx from taxonomy_data.ENTRIES."""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from taxonomy_data import ENTRIES

HEADERS = ["Main Category", "Subcategory", "Nested category", "Rank"]
WORKBOOK_NAME = "Course_Taxonomy_Master.xlsx"
DIST = ROOT / "dist"
PUBLIC = ROOT / "public"


def flatten_rows():
    rows = []
    for main, sub, topics in ENTRIES:
        for rank, nested in enumerate(topics, start=1):
            rows.append((main, sub, nested, rank))
    return rows


def sanitize_sheet_title(main: str) -> str:
    s = re.sub(r'[\[\]\*\/\\\?\:]', " ", main)
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 31:
        s = s[:31].rstrip()
    if not s:
        s = "Sheet"
    return s


def unique_sheet_names(mains: list[str]) -> dict[str, str]:
    used: set[str] = set()
    out: dict[str, str] = {}
    for main in mains:
        base = sanitize_sheet_title(main)
        name = base
        i = 1
        while name in used:
            suffix = f"_{i}"
            name = (base[: 31 - len(suffix)] + suffix).rstrip("_")
            i += 1
        used.add(name)
        out[main] = name
    return out


def write_sheet(ws, rows: list[tuple], table_display_name: str):
    ws.append(HEADERS)
    for r in rows:
        ws.append(list(r))
    n = len(rows) + 1
    ref = f"A1:{get_column_letter(len(HEADERS))}{n}"
    tab = Table(displayName=table_display_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        maxlen = max(len(str(c[col - 1])) for c in [HEADERS] + [list(x) for x in rows]) if rows else 10
        ws.column_dimensions[letter].width = min(52, max(12, maxlen + 2))


def export_csv(rows: list[tuple], path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(rows)


def main():
    rows = flatten_rows()
    mains_ordered = list(dict.fromkeys(m[0] for m in ENTRIES))
    name_map = unique_sheet_names(mains_ordered)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Master
    ws_m = wb.create_sheet("Master", 0)
    write_sheet(ws_m, rows, "tblMaster")

    # Per main category
    by_main: dict[str, list[tuple]] = defaultdict(list)
    for t in rows:
        by_main[t[0]].append(t)

    used_table_names: set[str] = {"tblMaster"}
    for main in mains_ordered:
        part = re.sub(r"[^0-9a-zA-Z]", "", main)[:20] or "Cat"
        tname = f"tbl_{part}"
        base = tname
        j = 1
        while tname in used_table_names:
            tname = f"{base[:24]}_{j}"
            j += 1
        used_table_names.add(tname)
        title = name_map[main]
        ws = wb.create_sheet(title)
        write_sheet(ws, by_main[main], tname)

    DIST.mkdir(parents=True, exist_ok=True)
    out_xlsx = DIST / WORKBOOK_NAME
    wb.save(out_xlsx)

    counts = Counter(r[0] for r in rows)
    sub_keys = {(r[0], r[1]) for r in rows}
    built = datetime.now(timezone.utc).isoformat()
    manifest = {
        "workbook": WORKBOOK_NAME,
        "built_at": built,
        "row_count": len(rows),
        "main_category_count": len(counts),
        "subcategory_count": len(sub_keys),
        "main_category_row_counts": dict(counts),
    }
    manifest_path = DIST / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    csv_path = DIST / "taxonomy.csv"
    export_csv(rows, csv_path)

    taxonomy_json = [
        {
            "main_category": r[0],
            "subcategory": r[1],
            "nested_category": r[2],
            "rank": r[3],
        }
        for r in rows
    ]
    json_path = DIST / "taxonomy.json"
    json_path.write_text(json.dumps(taxonomy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    PUBLIC.mkdir(parents=True, exist_ok=True)
    (PUBLIC / WORKBOOK_NAME).write_bytes(out_xlsx.read_bytes())
    (PUBLIC / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    (PUBLIC / "taxonomy.json").write_text(json.dumps(taxonomy_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (PUBLIC / "taxonomy.csv").write_text(csv_path.read_text(encoding="utf-8"), encoding="utf-8")

    print(f"Wrote {out_xlsx}")
    print(f"Wrote {manifest_path}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {json_path}")
    print(f"Copied workbook, manifest, taxonomy.json, taxonomy.csv to public/")


if __name__ == "__main__":
    main()
